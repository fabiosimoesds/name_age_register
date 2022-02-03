import pandas as pd
from openpyxl import load_workbook
from espfunc import *
wb = load_workbook('list_form.xlsx')
ws = wb.worksheets[0]


#MAIN MENU
while True:
    title('MAIN MENU')
    options('See list of People,Add new People,Delete a Person,Exit')
    opti = readInt('Choose your Option: ')


#READ THE LIST
    if opti == 1:
        read = pd.read_excel('list_form.xlsx', usecols='B, C').sort_values(0, ascending=True)
        list_name = read[0].tolist()
        list_age = read[1].tolist()
        title('List of People')
        for i, v in enumerate(list_name):
            print(f'{list_name[i]:<36} {list_age[i]:<3} anos')



#NEW DATA ENTRY
    elif opti == 2:
        data = []
        name = readName('Name: ')
        age = readInt('Age: ')
        data.append(name)
        data.append(age)
        print(f'{name} was added to the DataBase')
        ws[pers()] = data[0]
        ws[birt()] = data[1]
        wb.save('list_form.xlsx')
        data.clear()
    elif opti == 3:
        delet = readName('Type the name of the person you would like to delete: ')
        read = pd.read_excel('list_form.xlsx', usecols='B')[0].str.find(delet)
        for i, v in enumerate(read):
            if v == 0:
                rw = i + 2

        ws.delete_rows(rw)
        wb.save('list_form.xlsx')
    elif opti == 4:
        break
    else:
        line('+', 40)
        print(f'\'{opti}\' is not an option please chose again')
        line('+', 40)
